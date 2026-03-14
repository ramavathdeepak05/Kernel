"""
ALIS Mock API Server
Reads Admissions_Mockdata_final.xlsx and serves all frontend-facing endpoints.
Run: python mock_server.py
"""
import uuid, random, json, math, os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

# ─── Config ──────────────────────────────────────────────────────────────────
EXCEL_PATH = Path(__file__).parent / "Admissions_Mockdata_final.xlsx"
PORT = 8000
ADMIN_EMAIL = "admin@demo.edu"
ADMIN_PASSWORD = "Admin@1234"
MOCK_TOKEN = "mock-dev-token-alis-2025"

# Status distribution across 500 applicants (deterministic via seeded random)
STATUS_DIST = [
    ("DRAFT", 45),
    ("SUBMITTED", 40),
    ("UNDER_REVIEW", 30),
    ("DOCUMENTS_PENDING", 35),
    ("DOCUMENTS_RECEIVED", 25),
    ("ELIGIBILITY_CHECK", 20),
    ("ELIGIBLE", 45),
    ("INELIGIBLE", 15),
    ("TEST_SCHEDULED", 10),
    ("TEST_COMPLETED", 35),
    ("INTERVIEW_SCHEDULED", 8),
    ("INTERVIEW_COMPLETED", 27),
    ("MERIT_EVALUATED", 30),
    ("OFFER_ISSUED", 35),
    ("OFFER_ACCEPTED", 25),
    ("PAYMENT_PENDING", 10),
    ("PAYMENT_RECEIVED", 20),
    ("VERIFICATION_PENDING", 8),
    ("VERIFICATION_COMPLETE", 12),
    ("ENROLLED", 20),
    ("REJECTED", 18),
    ("WITHDRAWN", 7),
]

app = FastAPI(title="ALIS Mock API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Load Excel ───────────────────────────────────────────────────────────────
def load_sheet(wb, name: str) -> list[dict]:
    ws = wb[name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows

print("Loading Excel mock data...")
wb = openpyxl.load_workbook(str(EXCEL_PATH))
RAW_APPLICANTS   = load_sheet(wb, "Applicants")
RAW_ACADEMIC     = load_sheet(wb, "Academic_Qualifications")
RAW_EXAM         = load_sheet(wb, "Entrance_Exam_Scores")
RAW_PROGRAMS     = load_sheet(wb, "Program_Preferences")
RAW_DOCS         = load_sheet(wb, "Application_Documents")
print(f"Loaded {len(RAW_APPLICANTS)} applicants")

# ─── Build in-memory DB ───────────────────────────────────────────────────────
rng = random.Random(42)

# Expand status list
statuses = []
for status, count in STATUS_DIST:
    statuses.extend([status] * count)
rng.shuffle(statuses)
# Pad/trim to exactly len(RAW_APPLICANTS)
while len(statuses) < len(RAW_APPLICANTS):
    statuses.append("SUBMITTED")
statuses = statuses[:len(RAW_APPLICANTS)]

# Build lookup maps
academic_by_id: dict[str, list] = {}
for row in RAW_ACADEMIC:
    academic_by_id.setdefault(row["applicant_id"], []).append(row)

exam_by_id: dict[str, dict] = {r["applicant_id"]: r for r in RAW_EXAM}
prefs_by_id: dict[str, list] = {}
for row in RAW_PROGRAMS:
    prefs_by_id.setdefault(row["applicant_id"], []).append(row)
docs_by_id: dict[str, list] = {}
for row in RAW_DOCS:
    docs_by_id.setdefault(row["applicant_id"], []).append(row)

def make_uuid(seed: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"alis.{seed}"))

def fmt_date(d) -> str:
    if d is None:
        return None
    if isinstance(d, datetime):
        return d.isoformat()
    return str(d)

# Master applications store
APPLICATIONS: dict[str, dict] = {}
LEADS: dict[str, dict] = {}

now = datetime(2025, 3, 14, 10, 0, 0)

for i, raw in enumerate(RAW_APPLICANTS):
    app_id = make_uuid(raw["applicant_id"])
    lead_id = make_uuid("lead." + raw["applicant_id"])
    status = statuses[i]

    created_at = now - timedelta(days=rng.randint(1, 120))
    submitted_at = created_at + timedelta(days=rng.randint(0, 3)) if status != "DRAFT" else None

    exam = exam_by_id.get(raw["applicant_id"], {})
    prefs = prefs_by_id.get(raw["applicant_id"], [])
    programme = raw.get("intended_program") or (prefs[0]["program_name"] if prefs else "B.Tech")

    application = {
        "id": app_id,
        "application_id": raw["applicant_id"].replace("APP", "APP-2025-0"),
        "lead_id": lead_id,
        "applicant_name": raw["name"],
        "email": raw["email"],
        "phone": str(raw.get("phone") or ""),
        "programme": programme,
        "course": raw.get("course", ""),
        "major": raw.get("major", ""),
        "status": status,
        "step": 10 if status not in ("DRAFT", "SUBMITTED") else (1 if status == "DRAFT" else 10),
        "category": raw.get("category", "GENERAL"),
        "gender": raw.get("gender", ""),
        "date_of_birth": fmt_date(raw.get("date_of_birth")),
        "nationality": raw.get("nationality", "Indian"),
        "intake_batch": raw.get("intake_batch", "July-2025"),
        "study_mode": raw.get("study_mode", "FULL_TIME"),
        "hostel_required": raw.get("hostel_required", False),
        "scholarship_required": raw.get("scholarship_required", False),
        "has_disability": raw.get("has_disability", False),
        "work_experience_months": raw.get("work_experience_months", 0),
        "exam_name": exam.get("exam_name", ""),
        "exam_score": exam.get("score"),
        "exam_percentile": exam.get("percentile"),
        "exam_rank": exam.get("rank"),
        "form_data": {
            "first_name": raw["name"].split()[0],
            "last_name": " ".join(raw["name"].split()[1:]),
            "email": raw["email"],
            "programme": programme,
        },
        "submitted_at": fmt_date(submitted_at),
        "created_at": fmt_date(created_at),
        "updated_at": fmt_date(created_at + timedelta(hours=rng.randint(1, 48))),
    }

    # Lead (from applicant)
    lead_status_map = {
        "DRAFT": "READY_TO_APPLY",
        "SUBMITTED": "CONVERTED",
    }
    lead = {
        "id": lead_id,
        "name": raw["name"],
        "email": raw["email"],
        "phone": str(raw.get("phone") or ""),
        "status": lead_status_map.get(status, "CONVERTED"),
        "programme_interest": programme,
        "source": rng.choice(["Website", "Referral", "Social Media", "Walk-in", "Education Fair"]),
        "created_at": fmt_date(created_at - timedelta(days=rng.randint(1, 30))),
        "updated_at": fmt_date(created_at),
    }

    APPLICATIONS[app_id] = application
    LEADS[lead_id] = lead

APP_LIST = list(APPLICATIONS.values())
LEAD_LIST = list(LEADS.values())

print(f"Built {len(APP_LIST)} applications, {len(LEAD_LIST)} leads")

# ─── Pipeline summary ─────────────────────────────────────────────────────────
def get_pipeline_summary():
    counts: dict[str, int] = {}
    for a in APP_LIST:
        counts[a["status"]] = counts.get(a["status"], 0) + 1

    stages = [
        {"stage": "leads",        "label": "Leads",          "count": len(LEAD_LIST),                                        "color": "#a78bfa"},
        {"stage": "applications", "label": "Applications",   "count": counts.get("SUBMITTED", 0) + counts.get("UNDER_REVIEW", 0) + counts.get("DRAFT", 0), "color": "#818cf8"},
        {"stage": "documents",    "label": "Documents",      "count": counts.get("DOCUMENTS_PENDING", 0) + counts.get("DOCUMENTS_RECEIVED", 0), "color": "#60a5fa"},
        {"stage": "eligibility",  "label": "Eligibility",    "count": counts.get("ELIGIBLE", 0) + counts.get("INELIGIBLE", 0) + counts.get("ELIGIBILITY_CHECK", 0), "color": "#34d399"},
        {"stage": "tests",        "label": "Tests",          "count": counts.get("TEST_SCHEDULED", 0) + counts.get("TEST_COMPLETED", 0), "color": "#fbbf24"},
        {"stage": "merit",        "label": "Merit List",     "count": counts.get("MERIT_EVALUATED", 0),                      "color": "#f97316"},
        {"stage": "offers",       "label": "Offers",         "count": counts.get("OFFER_ISSUED", 0) + counts.get("OFFER_ACCEPTED", 0), "color": "#e879f9"},
        {"stage": "payment",      "label": "Payment",        "count": counts.get("PAYMENT_PENDING", 0) + counts.get("PAYMENT_RECEIVED", 0), "color": "#34d399"},
        {"stage": "verification", "label": "Verification",   "count": counts.get("VERIFICATION_PENDING", 0) + counts.get("VERIFICATION_COMPLETE", 0), "color": "#818cf8"},
        {"stage": "enrolled",     "label": "Enrolled",       "count": counts.get("ENROLLED", 0),                             "color": "#34d399"},
    ]
    return stages

# ─── Review queue ─────────────────────────────────────────────────────────────
def get_review_queue():
    items = []
    actionable = [a for a in APP_LIST if a["status"] in (
        "OFFER_ISSUED", "DOCUMENTS_PENDING", "ELIGIBILITY_CHECK",
        "VERIFICATION_PENDING", "PAYMENT_PENDING",
    )]
    for a in actionable[:20]:
        confidence = rng.randint(72, 96)
        urgency = "high" if confidence > 88 else "medium" if confidence > 78 else "low"
        type_map = {
            "OFFER_ISSUED":         ("ai_action",  f"Approve Conditional Offer: {a['applicant_name']}"),
            "DOCUMENTS_PENDING":    ("alert",      f"Documents Incomplete: {a['applicant_name']}"),
            "ELIGIBILITY_CHECK":    ("ai_action",  f"Borderline Eligibility: {a['applicant_name']}"),
            "VERIFICATION_PENDING": ("ai_action",  f"Final Verification Pending: {a['applicant_name']}"),
            "PAYMENT_PENDING":      ("exception",  f"Payment Timeout: {a['applicant_name']}"),
        }
        t, title = type_map.get(a["status"], ("alert", a["applicant_name"]))
        items.append({
            "id": make_uuid("review." + a["id"]),
            "type": t,
            "title": title,
            "subtitle": f"{a['programme']} · {a['application_id']}",
            "confidence": confidence,
            "urgency": urgency,
            "application_id": a["id"],
            "created_at": a["updated_at"],
        })
    return items

# ─── Offers ───────────────────────────────────────────────────────────────────
def build_offers():
    offers = []
    for a in APP_LIST:
        if a["status"] in ("OFFER_ISSUED", "OFFER_ACCEPTED", "OFFER_DECLINED",
                           "PAYMENT_PENDING", "PAYMENT_RECEIVED",
                           "VERIFICATION_PENDING", "VERIFICATION_COMPLETE", "ENROLLED"):
            offer_status_map = {
                "OFFER_ISSUED": "ISSUED",
                "OFFER_ACCEPTED": "ACCEPTED",
                "OFFER_DECLINED": "DECLINED",
                "PAYMENT_PENDING": "ACCEPTED",
                "PAYMENT_RECEIVED": "ACCEPTED",
                "VERIFICATION_PENDING": "COUNTERSIGNED",
                "VERIFICATION_COMPLETE": "COUNTERSIGNED",
                "ENROLLED": "COUNTERSIGNED",
            }
            issued = datetime.fromisoformat(a["updated_at"]) if a["updated_at"] else now
            offers.append({
                "id": make_uuid("offer." + a["id"]),
                "application_id": a["id"],
                "applicant_name": a["applicant_name"],
                "programme": a["programme"],
                "batch": a["intake_batch"],
                "fee_amount": str(rng.choice([85000, 95000, 110000, 125000, 150000])),
                "issued_at": issued.isoformat(),
                "expires_at": (issued + timedelta(days=15)).isoformat(),
                "status": offer_status_map.get(a["status"], "ISSUED"),
            })
    return offers

OFFERS = build_offers()

# ─── Payments ─────────────────────────────────────────────────────────────────
def build_payments():
    payments = []
    for a in APP_LIST:
        if a["status"] in ("PAYMENT_PENDING", "PAYMENT_RECEIVED",
                           "VERIFICATION_PENDING", "VERIFICATION_COMPLETE", "ENROLLED"):
            p_status = "SUCCESS" if a["status"] in (
                "PAYMENT_RECEIVED", "VERIFICATION_PENDING",
                "VERIFICATION_COMPLETE", "ENROLLED"
            ) else "PENDING"
            payments.append({
                "id": make_uuid("pay." + a["id"]),
                "application_id": a["id"],
                "amount": str(rng.choice([85000, 95000, 110000])),
                "currency": "INR",
                "status": p_status,
                "gateway": "Razorpay",
                "gateway_ref": f"pay_{rng.randint(100000, 999999)}" if p_status == "SUCCESS" else None,
                "paid_at": a["updated_at"] if p_status == "SUCCESS" else None,
                "created_at": a["updated_at"],
            })
    return payments

PAYMENTS = build_payments()

# ─── Enrollments ──────────────────────────────────────────────────────────────
def build_enrollments():
    enrollments = []
    for a in APP_LIST:
        if a["status"] == "ENROLLED":
            enrollments.append({
                "id": make_uuid("enroll." + a["id"]),
                "application_id": a["id"],
                "student_id": f"STU-2025-{rng.randint(10000, 99999)}",
                "programme": a["programme"],
                "batch": a["intake_batch"],
                "lms_provisioned": True,
                "hostel_allocated": a["hostel_required"],
                "library_id": f"LIB-{rng.randint(10000, 99999)}",
                "enrolled_at": a["updated_at"],
            })
    return enrollments

ENROLLMENTS = build_enrollments()

# ─── Seat matrix ──────────────────────────────────────────────────────────────
PROGRAMS = ["B.Tech - CSE", "B.Tech - ECE", "B.Tech - ME", "M.Tech - Data Science",
            "M.Tech - AI", "MBA", "BBA", "MCA", "B.Sc - Physics"]
CATEGORIES = ["GENERAL", "SC", "ST", "OBC_NCL", "EWS"]

SEAT_MATRIX = []
for prog in PROGRAMS:
    total = rng.randint(40, 120)
    filled = rng.randint(10, total)
    SEAT_MATRIX.append({
        "id": make_uuid(f"seat.{prog}.GENERAL"),
        "programme": prog,
        "category": "GENERAL",
        "total_seats": total,
        "filled_seats": filled,
    })

# ─── Documents per application ────────────────────────────────────────────────
DOC_STATUS_MAP = {
    "DRAFT": "PENDING",
    "SUBMITTED": "PENDING",
    "UNDER_REVIEW": "UNDER_REVIEW",
    "DOCUMENTS_PENDING": "PENDING",
    "DOCUMENTS_RECEIVED": "APPROVED",
    "ELIGIBLE": "APPROVED",
    "ENROLLED": "APPROVED",
}

# ─── Auth ─────────────────────────────────────────────────────────────────────
ADMIN_USER = {
    "id": make_uuid("admin"),
    "name": "System Administrator",
    "email": ADMIN_EMAIL,
    "role": "SUPER_ADMIN",
    "tenant_id": make_uuid("demo-org"),
    "status": "ACTIVE",
}

def require_auth(request: Request):
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    # Accept any non-empty token in dev mode
    return ADMIN_USER

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "mode": "mock", "applicants": len(APP_LIST)}

# Auth
@app.post("/api/v1/auth/login")
async def login(request: Request):
    body = await request.json()
    if body.get("email") == ADMIN_EMAIL and body.get("password") == ADMIN_PASSWORD:
        return {"access_token": MOCK_TOKEN, "token_type": "bearer", "user": ADMIN_USER}
    raise HTTPException(status_code=401, detail="Invalid credentials")

@app.get("/api/v1/auth/me")
async def me(request: Request):
    require_auth(request)
    return ADMIN_USER

@app.post("/api/v1/auth/logout")
async def logout():
    return {"detail": "Logged out"}

# ─── Admissions: pipeline + review queue ─────────────────────────────────────

@app.get("/api/v1/admissions/pipeline-summary")
async def pipeline_summary(request: Request):
    require_auth(request)
    return get_pipeline_summary()

@app.get("/api/v1/admissions/review-queue")
async def review_queue(request: Request):
    require_auth(request)
    return get_review_queue()

# ─── Applicants (legacy) ─────────────────────────────────────────────────────

@app.get("/api/v1/admissions/applicants")
async def list_applicants(request: Request, limit: int = 50, skip: int = 0):
    require_auth(request)
    return APP_LIST[skip: skip + limit]

# ─── Leads ───────────────────────────────────────────────────────────────────

@app.get("/api/v1/admissions/leads")
async def list_leads(request: Request, skip: int = 0, limit: int = 50, status: str = None):
    require_auth(request)
    items = LEAD_LIST
    if status:
        items = [l for l in items if l["status"] == status]
    return items[skip: skip + limit]

@app.get("/api/v1/admissions/leads/{lead_id}")
async def get_lead(lead_id: str, request: Request):
    require_auth(request)
    if lead_id not in LEADS:
        raise HTTPException(404, "Lead not found")
    return LEADS[lead_id]

@app.post("/api/v1/admissions/leads")
async def create_lead(request: Request):
    require_auth(request)
    body = await request.json()
    new_id = str(uuid.uuid4())
    lead = {"id": new_id, "status": "NEW", "created_at": now.isoformat(), "updated_at": now.isoformat(), **body}
    LEADS[new_id] = lead
    LEAD_LIST.append(lead)
    return lead

@app.post("/api/v1/admissions/leads/{lead_id}/transition")
async def transition_lead(lead_id: str, request: Request):
    require_auth(request)
    body = await request.json()
    if lead_id not in LEADS:
        raise HTTPException(404, "Lead not found")
    LEADS[lead_id]["status"] = body.get("new_status", LEADS[lead_id]["status"])
    return LEADS[lead_id]

@app.get("/api/v1/admissions/leads/{lead_id}/activities")
async def lead_activities(lead_id: str, request: Request):
    require_auth(request)
    return []

@app.post("/api/v1/admissions/leads/{lead_id}/activities")
async def add_lead_activity(lead_id: str, request: Request):
    require_auth(request)
    body = await request.json()
    return {"id": str(uuid.uuid4()), "lead_id": lead_id, "created_at": now.isoformat(), **body}

@app.post("/api/v1/admissions/leads/{lead_id}/convert")
async def convert_lead(lead_id: str, request: Request):
    require_auth(request)
    if lead_id not in LEADS:
        raise HTTPException(404, "Lead not found")
    LEADS[lead_id]["status"] = "CONVERTED"
    return {"id": str(uuid.uuid4()), "status": "SUBMITTED", "lead_id": lead_id}

# ─── Applications ─────────────────────────────────────────────────────────────

@app.get("/api/v1/admissions/applications")
async def list_applications(request: Request, skip: int = 0, limit: int = 50,
                            status: str = None, program: str = None):
    require_auth(request)
    items = APP_LIST
    if status:
        items = [a for a in items if a["status"] == status]
    if program:
        items = [a for a in items if program.lower() in a["programme"].lower()]
    return items[skip: skip + limit]

@app.get("/api/v1/admissions/applications/{app_id}")
async def get_application(app_id: str, request: Request):
    require_auth(request)
    if app_id not in APPLICATIONS:
        raise HTTPException(404, "Application not found")
    return APPLICATIONS[app_id]

@app.post("/api/v1/admissions/applications")
async def create_application(request: Request):
    require_auth(request)
    body = await request.json()
    new_id = str(uuid.uuid4())
    app_record = {
        "id": new_id,
        "application_id": f"APP-2025-{rng.randint(100000, 999999)}",
        "status": "DRAFT",
        "step": 1,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        **body,
    }
    APPLICATIONS[new_id] = app_record
    APP_LIST.append(app_record)
    return app_record

@app.patch("/api/v1/admissions/applications/{app_id}")
async def update_application(app_id: str, request: Request):
    require_auth(request)
    if app_id not in APPLICATIONS:
        raise HTTPException(404, "Application not found")
    body = await request.json()
    APPLICATIONS[app_id].update(body)
    APPLICATIONS[app_id]["updated_at"] = now.isoformat()
    return APPLICATIONS[app_id]

@app.post("/api/v1/admissions/applications/{app_id}/submit")
async def submit_application(app_id: str, request: Request):
    require_auth(request)
    if app_id not in APPLICATIONS:
        raise HTTPException(404, "Application not found")
    APPLICATIONS[app_id]["status"] = "SUBMITTED"
    APPLICATIONS[app_id]["submitted_at"] = now.isoformat()
    return APPLICATIONS[app_id]

@app.post("/api/v1/admissions/applications/{app_id}/withdraw")
async def withdraw_application(app_id: str, request: Request):
    require_auth(request)
    if app_id not in APPLICATIONS:
        raise HTTPException(404, "Application not found")
    APPLICATIONS[app_id]["status"] = "WITHDRAWN"
    return APPLICATIONS[app_id]

# ─── Documents ────────────────────────────────────────────────────────────────

@app.get("/api/v1/admissions/applications/{app_id}/documents")
async def list_documents(app_id: str, request: Request):
    require_auth(request)
    app_record = APPLICATIONS.get(app_id)
    if not app_record:
        raise HTTPException(404, "Application not found")
    doc_status = DOC_STATUS_MAP.get(app_record["status"], "APPROVED")
    doc_types = ["PHOTO", "ID_PROOF", "CLASS_10_MARKSHEET", "CLASS_12_MARKSHEET",
                 "ENTRANCE_SCORE_CARD", "CATEGORY_CERTIFICATE"]
    return [
        {
            "id": make_uuid(f"doc.{app_id}.{dt}"),
            "application_id": app_id,
            "type": dt,
            "status": doc_status,
            "file_url": f"https://mock-bucket/docs/{app_id}/{dt.lower()}.pdf",
            "uploaded_at": app_record.get("updated_at"),
        }
        for dt in doc_types
    ]

@app.post("/api/v1/admissions/applications/{app_id}/documents/upload")
async def upload_document(app_id: str, request: Request):
    require_auth(request)
    return {"id": str(uuid.uuid4()), "application_id": app_id, "status": "PENDING",
            "uploaded_at": now.isoformat()}

@app.post("/api/v1/admissions/applications/{app_id}/documents/{doc_id}/review")
async def review_document(app_id: str, doc_id: str, request: Request):
    require_auth(request)
    body = await request.json()
    return {"id": doc_id, "application_id": app_id, "status": body.get("status", "APPROVED"),
            "reviewed_at": now.isoformat()}

@app.get("/api/v1/admissions/documents/review-queue")
async def doc_review_queue(request: Request, skip: int = 0, limit: int = 20):
    require_auth(request)
    pending = [a for a in APP_LIST if a["status"] == "DOCUMENTS_PENDING"]
    return [
        {"id": make_uuid(f"doc.{a['id']}.review"), "application_id": a["id"],
         "type": "CLASS_12_MARKSHEET", "status": "UNDER_REVIEW",
         "uploaded_at": a["updated_at"]}
        for a in pending[skip: skip + limit]
    ]

# ─── Eligibility ─────────────────────────────────────────────────────────────

@app.post("/api/v1/admissions/eligibility/check/{app_id}")
async def check_eligibility(app_id: str, request: Request):
    require_auth(request)
    app_record = APPLICATIONS.get(app_id)
    if not app_record:
        raise HTTPException(404, "Application not found")
    is_eligible = app_record["status"] not in ("INELIGIBLE", "REJECTED")
    return {
        "application_id": app_id,
        "is_eligible": is_eligible,
        "checked_at": now.isoformat(),
        "criteria": [
            {"label": "12th percentage >= 50%", "passed": is_eligible, "detail": "63.5%"},
            {"label": "Valid entrance score", "passed": True, "detail": "JEE Mains"},
            {"label": "Age criteria", "passed": True, "detail": "18-25 years"},
        ],
    }

@app.get("/api/v1/admissions/eligibility/{app_id}")
async def get_eligibility(app_id: str, request: Request):
    require_auth(request)
    app_record = APPLICATIONS.get(app_id)
    if not app_record:
        raise HTTPException(404, "Application not found")
    return {
        "application_id": app_id,
        "is_eligible": app_record["status"] not in ("INELIGIBLE", "REJECTED"),
        "checked_at": app_record.get("updated_at", now.isoformat()),
        "criteria": [],
    }

@app.post("/api/v1/admissions/eligibility/{app_id}/override")
async def override_eligibility(app_id: str, request: Request):
    require_auth(request)
    body = await request.json()
    return {"application_id": app_id, "is_eligible": body.get("verdict") == "ELIGIBLE",
            "override": True, "override_reason": body.get("override_reason"),
            "checked_at": now.isoformat(), "criteria": []}

@app.post("/api/v1/admissions/eligibility/batch-check")
async def batch_eligibility(request: Request):
    require_auth(request)
    body = await request.json()
    return [{"application_id": aid, "is_eligible": True, "checked_at": now.isoformat(), "criteria": []}
            for aid in body.get("application_ids", [])]

# ─── Entrance Tests ──────────────────────────────────────────────────────────

MOCK_TESTS = [
    {"id": make_uuid("test.jee"), "name": "JEE Mains (Internal)", "programme": "B.Tech",
     "date": "2025-02-15", "duration_mins": 180, "max_score": 300, "status": "COMPLETED"},
    {"id": make_uuid("test.cat"), "name": "CAT Score Verification", "programme": "MBA",
     "date": "2025-01-20", "duration_mins": 120, "max_score": 100, "status": "COMPLETED"},
    {"id": make_uuid("test.mtech"), "name": "GATE / Internal Test", "programme": "M.Tech",
     "date": "2025-03-10", "duration_mins": 180, "max_score": 100, "status": "COMPLETED"},
]

@app.get("/api/v1/admissions/entrance-tests")
async def list_tests(request: Request):
    require_auth(request)
    return MOCK_TESTS

@app.get("/api/v1/admissions/entrance-tests/{test_id}")
async def get_test(test_id: str, request: Request):
    require_auth(request)
    for t in MOCK_TESTS:
        if t["id"] == test_id:
            return t
    raise HTTPException(404, "Test not found")

@app.post("/api/v1/admissions/entrance-tests")
async def create_test(request: Request):
    require_auth(request)
    body = await request.json()
    new = {"id": str(uuid.uuid4()), "status": "SCHEDULED", **body}
    MOCK_TESTS.append(new)
    return new

@app.get("/api/v1/admissions/entrance-tests/{test_id}/slots")
async def test_slots(test_id: str, request: Request):
    require_auth(request)
    return [{"id": make_uuid(f"slot.{test_id}.1"), "test_id": test_id,
             "venue": "Main Examination Hall", "start_time": "2025-02-15T09:00:00",
             "capacity": 200, "registered": 178}]

@app.post("/api/v1/admissions/entrance-tests/{test_id}/register")
async def register_test(test_id: str, request: Request):
    require_auth(request)
    body = await request.json()
    return {"id": str(uuid.uuid4()), "test_id": test_id, "status": "REGISTERED", **body}

@app.get("/api/v1/admissions/entrance-tests/{test_id}/registrations")
async def test_registrations(test_id: str, request: Request):
    require_auth(request)
    test_apps = [a for a in APP_LIST if a["status"] in ("TEST_SCHEDULED", "TEST_COMPLETED")][:30]
    return [{"id": make_uuid(f"reg.{test_id}.{a['id']}"), "application_id": a["id"],
             "slot_id": make_uuid(f"slot.{test_id}.1"), "status": "APPEARED",
             "score": a.get("exam_score"), "rank": a.get("exam_rank")}
            for a in test_apps]

@app.post("/api/v1/admissions/entrance-tests/{test_id}/scores")
async def record_score(test_id: str, request: Request):
    require_auth(request)
    body = await request.json()
    return {"id": str(uuid.uuid4()), "test_id": test_id, "status": "APPEARED", **body}

@app.post("/api/v1/admissions/entrance-tests/{test_id}/admit-cards")
async def gen_admit_cards(test_id: str, request: Request):
    require_auth(request)
    return {"generated": 178}

# ─── Interviews ──────────────────────────────────────────────────────────────

MOCK_INTERVIEWS = []
interview_apps = [a for a in APP_LIST if a["status"] in ("INTERVIEW_SCHEDULED", "INTERVIEW_COMPLETED")][:20]
for a in interview_apps:
    MOCK_INTERVIEWS.append({
        "id": make_uuid(f"iv.{a['id']}"),
        "application_id": a["id"],
        "panel_id": make_uuid("panel.1"),
        "scheduled_at": "2025-02-20T10:00:00",
        "mode": "IN_PERSON",
        "score": rng.randint(55, 90) if a["status"] == "INTERVIEW_COMPLETED" else None,
        "status": "COMPLETED" if a["status"] == "INTERVIEW_COMPLETED" else "SCHEDULED",
    })

@app.get("/api/v1/admissions/interview-panels")
async def interview_panels(request: Request):
    require_auth(request)
    return [{"id": make_uuid("panel.1"), "name": "Panel A — Engineering", "program": "B.Tech"},
            {"id": make_uuid("panel.2"), "name": "Panel B — Management", "program": "MBA"}]

@app.get("/api/v1/admissions/interviews")
async def list_interviews(request: Request, status: str = None):
    require_auth(request)
    items = MOCK_INTERVIEWS
    if status:
        items = [i for i in items if i["status"] == status]
    return items

@app.get("/api/v1/admissions/interviews/{iv_id}")
async def get_interview(iv_id: str, request: Request):
    require_auth(request)
    for i in MOCK_INTERVIEWS:
        if i["id"] == iv_id:
            return i
    raise HTTPException(404, "Interview not found")

@app.post("/api/v1/admissions/interviews")
async def schedule_interview(request: Request):
    require_auth(request)
    body = await request.json()
    new = {"id": str(uuid.uuid4()), "status": "SCHEDULED", **body}
    MOCK_INTERVIEWS.append(new)
    return new

@app.post("/api/v1/admissions/interviews/{iv_id}/scorecard")
async def submit_scorecard(iv_id: str, request: Request):
    require_auth(request)
    body = await request.json()
    for i in MOCK_INTERVIEWS:
        if i["id"] == iv_id:
            i.update(body)
            return i
    raise HTTPException(404)

@app.post("/api/v1/admissions/interviews/{iv_id}/complete")
async def complete_interview(iv_id: str, request: Request):
    require_auth(request)
    for i in MOCK_INTERVIEWS:
        if i["id"] == iv_id:
            i["status"] = "COMPLETED"
            return i
    raise HTTPException(404)

@app.post("/api/v1/admissions/interviews/{iv_id}/no-show")
async def no_show_interview(iv_id: str, request: Request):
    require_auth(request)
    for i in MOCK_INTERVIEWS:
        if i["id"] == iv_id:
            i["status"] = "NO_SHOW"
            return i
    raise HTTPException(404)

# ─── Merit Lists ─────────────────────────────────────────────────────────────

merit_apps = [a for a in APP_LIST if a["status"] == "MERIT_EVALUATED"]
MERIT_ENTRIES = sorted(
    [{"rank": i + 1, "application_id": a["id"], "applicant_name": a["applicant_name"],
      "score": round(rng.uniform(55, 98), 2), "category": a.get("category", "GENERAL"),
      "status": "SELECTED" if i < 20 else ("WAITLISTED" if i < 28 else "REJECTED")}
     for i, a in enumerate(merit_apps)],
    key=lambda x: x["rank"]
)

MERIT_LISTS = [{
    "id": make_uuid("merit.btech.2025"),
    "programme": "B.Tech",
    "batch": "July-2025",
    "status": "PUBLISHED",
    "generated_at": "2025-03-01T10:00:00",
    "entries": MERIT_ENTRIES[:30],
}]

@app.get("/api/v1/admissions/merit-lists")
async def list_merit(request: Request):
    require_auth(request)
    return [{"id": m["id"], "programme": m["programme"], "batch": m["batch"],
             "status": m["status"], "generated_at": m["generated_at"]} for m in MERIT_LISTS]

@app.get("/api/v1/admissions/merit-lists/{ml_id}")
async def get_merit(ml_id: str, request: Request):
    require_auth(request)
    for m in MERIT_LISTS:
        if m["id"] == ml_id:
            return m
    raise HTTPException(404)

@app.post("/api/v1/admissions/merit-lists/generate")
async def generate_merit(request: Request):
    require_auth(request)
    body = await request.json()
    return {"id": str(uuid.uuid4()), "status": "DRAFT", "generated_at": now.isoformat(), **body, "entries": []}

@app.post("/api/v1/admissions/merit-lists/{ml_id}/approve")
async def approve_merit(ml_id: str, request: Request):
    require_auth(request)
    return {"id": ml_id, "status": "APPROVED"}

@app.post("/api/v1/admissions/merit-lists/{ml_id}/publish")
async def publish_merit(ml_id: str, request: Request):
    require_auth(request)
    return {"id": ml_id, "status": "PUBLISHED"}

@app.get("/api/v1/admissions/seat-matrix")
async def seat_matrix(request: Request):
    require_auth(request)
    return SEAT_MATRIX

# ─── Offer Letters ────────────────────────────────────────────────────────────

OFFERS_BY_ID = {o["id"]: o for o in OFFERS}

@app.get("/api/v1/admissions/offer-letters")
async def list_offers(request: Request, status: str = None):
    require_auth(request)
    items = OFFERS
    if status:
        items = [o for o in items if o["status"] == status]
    return items

@app.get("/api/v1/admissions/offer-letters/{offer_id}")
async def get_offer(offer_id: str, request: Request):
    require_auth(request)
    if offer_id not in OFFERS_BY_ID:
        raise HTTPException(404)
    return OFFERS_BY_ID[offer_id]

@app.get("/api/v1/admissions/offer-letters/application/{app_id}")
async def offer_by_application(app_id: str, request: Request):
    require_auth(request)
    for o in OFFERS:
        if o["application_id"] == app_id:
            return o
    raise HTTPException(404, "No offer for this application")

@app.post("/api/v1/admissions/offer-letters/generate")
async def generate_offer(request: Request):
    require_auth(request)
    body = await request.json()
    new = {"id": str(uuid.uuid4()), "status": "ISSUED",
           "issued_at": now.isoformat(),
           "expires_at": (now + timedelta(days=15)).isoformat(),
           "fee_amount": "95000", **body}
    OFFERS.append(new)
    OFFERS_BY_ID[new["id"]] = new
    return new

@app.post("/api/v1/admissions/offer-letters/batch-generate")
async def batch_generate_offers(request: Request):
    require_auth(request)
    body = await request.json()
    return {"generated": len(body.get("application_ids", []))}

@app.post("/api/v1/admissions/offer-letters/{offer_id}/accept")
async def accept_offer(offer_id: str, request: Request):
    require_auth(request)
    if offer_id in OFFERS_BY_ID:
        OFFERS_BY_ID[offer_id]["status"] = "ACCEPTED"
        return OFFERS_BY_ID[offer_id]
    raise HTTPException(404)

@app.post("/api/v1/admissions/offer-letters/{offer_id}/decline")
async def decline_offer(offer_id: str, request: Request):
    require_auth(request)
    if offer_id in OFFERS_BY_ID:
        OFFERS_BY_ID[offer_id]["status"] = "DECLINED"
        return OFFERS_BY_ID[offer_id]
    raise HTTPException(404)

@app.post("/api/v1/admissions/offer-letters/{offer_id}/countersign")
async def countersign_offer(offer_id: str, request: Request):
    require_auth(request)
    if offer_id in OFFERS_BY_ID:
        OFFERS_BY_ID[offer_id]["status"] = "COUNTERSIGNED"
        return OFFERS_BY_ID[offer_id]
    raise HTTPException(404)

# ─── Payments ─────────────────────────────────────────────────────────────────

PAY_BY_ID = {p["id"]: p for p in PAYMENTS}

@app.get("/api/v1/admissions/payments")
async def list_payments(request: Request, application_id: str = None, status: str = None):
    require_auth(request)
    items = PAYMENTS
    if application_id:
        items = [p for p in items if p["application_id"] == application_id]
    if status:
        items = [p for p in items if p["status"] == status]
    return items

@app.get("/api/v1/admissions/payments/{pay_id}")
async def get_payment(pay_id: str, request: Request):
    require_auth(request)
    if pay_id not in PAY_BY_ID:
        raise HTTPException(404)
    return PAY_BY_ID[pay_id]

@app.post("/api/v1/admissions/payments/initiate")
async def initiate_payment(request: Request):
    require_auth(request)
    body = await request.json()
    pay_id = str(uuid.uuid4())
    return {"payment_id": pay_id, "gateway_order_id": f"order_{rng.randint(100000,999999)}",
            "amount": "1000", "gateway_url": "https://mock-razorpay/checkout"}

@app.post("/api/v1/admissions/payments/{pay_id}/verify")
async def verify_payment(pay_id: str, request: Request):
    require_auth(request)
    return {"id": pay_id, "status": "SUCCESS", "paid_at": now.isoformat()}

@app.post("/api/v1/admissions/payments/{pay_id}/refund")
async def refund_payment(pay_id: str, request: Request):
    require_auth(request)
    if pay_id in PAY_BY_ID:
        PAY_BY_ID[pay_id]["status"] = "REFUNDED"
        return PAY_BY_ID[pay_id]
    raise HTTPException(404)

# ─── Final Verification ───────────────────────────────────────────────────────

@app.get("/api/v1/admissions/final-verification/{app_id}")
async def get_verification(app_id: str, request: Request):
    require_auth(request)
    app_record = APPLICATIONS.get(app_id)
    if not app_record:
        raise HTTPException(404)
    return {"id": make_uuid(f"fv.{app_id}"), "application_id": app_id,
            "status": "APPROVED" if app_record["status"] in ("VERIFICATION_COMPLETE", "ENROLLED") else "PENDING",
            "doc_check": True, "background_check": None}

@app.post("/api/v1/admissions/final-verification/{app_id}/initiate")
async def initiate_verification(app_id: str, request: Request):
    require_auth(request)
    return {"id": make_uuid(f"fv.{app_id}"), "application_id": app_id, "status": "IN_PROGRESS",
            "doc_check": False}

@app.post("/api/v1/admissions/final-verification/{app_id}/check-document")
async def check_verification_doc(app_id: str, request: Request):
    require_auth(request)
    return {"id": make_uuid(f"fv.{app_id}"), "application_id": app_id, "status": "IN_PROGRESS",
            "doc_check": True}

@app.post("/api/v1/admissions/final-verification/{app_id}/clear")
async def clear_verification(app_id: str, request: Request):
    require_auth(request)
    if app_id in APPLICATIONS:
        APPLICATIONS[app_id]["status"] = "VERIFICATION_COMPLETE"
    return {"id": make_uuid(f"fv.{app_id}"), "application_id": app_id, "status": "APPROVED",
            "doc_check": True}

# ─── Enrollment ──────────────────────────────────────────────────────────────

ENROLL_BY_APP = {e["application_id"]: e for e in ENROLLMENTS}

@app.get("/api/v1/admissions/enrollment")
async def list_enrollments(request: Request, status: str = None):
    require_auth(request)
    return ENROLLMENTS

@app.get("/api/v1/admissions/enrollment/{app_id}")
async def get_enrollment(app_id: str, request: Request):
    require_auth(request)
    if app_id not in ENROLL_BY_APP:
        raise HTTPException(404, "Not enrolled")
    return ENROLL_BY_APP[app_id]

@app.post("/api/v1/admissions/enrollment/{app_id}/provision")
async def provision_enrollment(app_id: str, request: Request):
    require_auth(request)
    if app_id in ENROLL_BY_APP:
        return ENROLL_BY_APP[app_id]
    if app_id not in APPLICATIONS:
        raise HTTPException(404)
    enrollment = {
        "id": str(uuid.uuid4()),
        "application_id": app_id,
        "student_id": f"STU-2025-{rng.randint(10000, 99999)}",
        "programme": APPLICATIONS[app_id].get("programme", ""),
        "batch": "July-2025",
        "lms_provisioned": True,
        "hostel_allocated": APPLICATIONS[app_id].get("hostel_required", False),
        "library_id": f"LIB-{rng.randint(10000, 99999)}",
        "enrolled_at": now.isoformat(),
    }
    APPLICATIONS[app_id]["status"] = "ENROLLED"
    ENROLLMENTS.append(enrollment)
    ENROLL_BY_APP[app_id] = enrollment
    return enrollment

@app.post("/api/v1/admissions/enrollment/batch-provision")
async def batch_provision(request: Request):
    require_auth(request)
    body = await request.json()
    return {"provisioned": len(body.get("application_ids", [])), "failed": 0}

# ─── Catch-all for any unimplemented routes ───────────────────────────────────
@app.api_route("/api/v1/{path:path}", methods=["GET","POST","PUT","PATCH","DELETE"])
async def catch_all(path: str, request: Request):
    return JSONResponse({"detail": f"Mock: /{path} not yet implemented", "mock": True}, status_code=200)

# ─── Run ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    print(f"\nALIS Mock API -> http://localhost:{PORT}")
    print(f"Login: {ADMIN_EMAIL} / {ADMIN_PASSWORD}")
    print(f"Docs:  http://localhost:{PORT}/docs\n")
    uvicorn.run(app, host="0.0.0.0", port=PORT, log_level="info")
