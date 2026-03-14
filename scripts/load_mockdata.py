"""
Load Admissions_Mockdata_final.xlsx into PostgreSQL for development.

Usage:
    cd ALIS
    python ../scripts/load_mockdata.py
"""
import sys
import os
import json
import random
import uuid
import psycopg2
import openpyxl
from datetime import datetime, timezone

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DB_URL = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/alis_db")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Admissions_Mockdata_final.xlsx")
ORG_ID = "demo"

STATUSES = [
    "NEW", "NEW", "NEW",            # 3× — most leads are new
    "CONTACTED", "CONTACTED",
    "INTERESTED",
    "READY_TO_APPLY",
    "CONVERTED", "CONVERTED",       # 2× — a good chunk convert
    "DISQUALIFIED",
]

SOURCE_TYPES = ["WEBSITE", "REFERRAL", "WALK_IN", "SOCIAL_MEDIA", "AGENT", "EVENT"]

# ---------------------------------------------------------------------------
# Load Excel
# ---------------------------------------------------------------------------
print("Loading Excel...", flush=True)
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Applicants"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

applicants = []
for row_idx in range(2, ws.max_row + 1):
    row = {headers[c - 1]: ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)}
    applicants.append(row)

print(f"  Loaded {len(applicants)} applicants from Excel", flush=True)

# ---------------------------------------------------------------------------
# Connect to DB
# ---------------------------------------------------------------------------
conn = psycopg2.connect(DB_URL)
conn.autocommit = False
cur = conn.cursor()

# Set tenant context for RLS
cur.execute("SET LOCAL alis.current_tenant = 'demo'")

# Count existing leads
cur.execute("SELECT COUNT(*) FROM leads WHERE org_id = %s", (ORG_ID,))
existing = cur.fetchone()[0]
if existing > 0:
    print(f"  {existing} leads already exist for org '{ORG_ID}' — skipping import.")
    conn.close()
    sys.exit(0)

# ---------------------------------------------------------------------------
# Insert leads
# ---------------------------------------------------------------------------
print("Inserting leads...", flush=True)
inserted = 0
rng = random.Random(42)

for app in applicants:
    name = app.get("name") or "Unknown"
    email = app.get("email") or f"unknown_{uuid.uuid4().hex[:8]}@demo.edu"
    phone = str(app.get("phone") or "")[:20]

    # Extract city/state from permanent_address JSON if available
    city, state = "", ""
    addr_raw = app.get("permanent_address")
    if addr_raw:
        try:
            addr = json.loads(addr_raw) if isinstance(addr_raw, str) else addr_raw
            city = addr.get("city", "")
            state = addr.get("state", "")
        except Exception:
            pass

    program = app.get("intended_program") or app.get("course") or "B.Tech"
    status = rng.choice(STATUSES)
    source_type = rng.choice(SOURCE_TYPES)
    intake_year = str(app.get("intake_batch") or "2025")[:10]

    try:
        cur.execute(
            """
            INSERT INTO leads (
                id, org_id, full_name, email, phone, city, state_region,
                course_interest, intake_year, source_type, status, created_at, updated_at
            ) VALUES (
                uuid_generate_v4(), %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, NOW(), NOW()
            )
            """,
            (
                ORG_ID, name, email, phone, city, state,
                program, intake_year, source_type, status,
            ),
        )
        inserted += 1
    except Exception as e:
        print(f"  Skipping row (email={email}): {e}", flush=True)
        conn.rollback()
        cur.execute("SET LOCAL alis.current_tenant = 'demo'")

conn.commit()
print(f"  Inserted {inserted} leads into 'leads' table.", flush=True)

# ---------------------------------------------------------------------------
# Insert into applicants table for converted leads
# ---------------------------------------------------------------------------
print("Inserting applicants for CONVERTED leads...", flush=True)

cur.execute("SET LOCAL alis.current_tenant = 'demo'")
cur.execute(
    "SELECT COUNT(*) FROM applicants WHERE org_id = %s", (ORG_ID,)
)
existing_apps = cur.fetchone()[0]

if existing_apps == 0:
    cur.execute(
        "SELECT id, full_name, email, phone, course_interest FROM leads WHERE org_id = %s AND status = 'CONVERTED'",
        (ORG_ID,),
    )
    converted_leads = cur.fetchall()
    app_inserted = 0
    for lead_id, name, email, phone, program in converted_leads:
        try:
            cur.execute(
                """
                INSERT INTO applicants (
                    id, org_id, name, email, phone, intended_program,
                    source_channel, status, created_at
                ) VALUES (
                    uuid_generate_v4(), %s, %s, %s, %s, %s, 'WEBSITE', 'APPLIED', NOW()
                )
                """,
                (ORG_ID, name, email, phone, program),
            )
            app_inserted += 1
        except Exception as e:
            conn.rollback()
            cur.execute("SET LOCAL alis.current_tenant = 'demo'")

    conn.commit()
    print(f"  Inserted {app_inserted} applicants.", flush=True)
else:
    print(f"  {existing_apps} applicants already exist — skipping.", flush=True)

cur.close()
conn.close()
print("Done!", flush=True)
